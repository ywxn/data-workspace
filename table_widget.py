"""
Interactive table widget with sorting, filtering, and export capabilities.

This module provides an enhanced QTableWidget with advanced features including
column-based sorting, search filtering, row selection, and export to CSV/Excel.
"""

import csv
from typing import List, Dict, Any, Optional

from PySide6.QtWidgets import (
    QTableWidget,
    QTableWidgetItem,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLineEdit,
    QLabel,
    QHeaderView,
    QFileDialog,
    QMenu,
    QApplication,
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QAction


class InteractiveTableWidget(QWidget):
    """Interactive table with sorting, filtering, and export capabilities."""

    row_selected = Signal(int)  # Emitted when a row is selected

    def __init__(self, parent=None):
        super().__init__(parent)
        self.original_data: List[List[Any]] = []
        self.headers: List[str] = []
        self.table = None
        self.search_input = None
        self.clear_filter_btn = None
        self.copy_btn = None
        self.export_btn = None
        self.row_count_label = None
        self.setup_ui()

    def setup_ui(self):
        """Initialize UI components."""
        layout = QVBoxLayout(self)

        # Search/filter bar
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Type to filter rows...")
        self.search_input.textChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.search_input)
        self.clear_filter_btn = QPushButton("Clear")
        self.clear_filter_btn.clicked.connect(self.clear_filter)
        filter_layout.addWidget(self.clear_filter_btn)
        layout.addLayout(filter_layout)

        # Table widget
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        # Enable auto-resize for better UX
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self.table.horizontalHeader().setStretchLastSection(True)

        layout.addWidget(self.table)

        # Status and action bar
        bottom_layout = QHBoxLayout()
        self.row_count_label = QLabel("Rows: 0")
        bottom_layout.addWidget(self.row_count_label)
        bottom_layout.addStretch()

        self.copy_btn = QPushButton("Copy Selected")
        self.copy_btn.clicked.connect(self.copy_selected)
        bottom_layout.addWidget(self.copy_btn)

        self.export_btn = QPushButton("Export to CSV")
        self.export_btn.clicked.connect(self.export_to_csv)
        bottom_layout.addWidget(self.export_btn)

        layout.addLayout(bottom_layout)

    def load_data(self, headers: List[str], rows: List[List[Any]]):
        """
        Load data into the table.

        Args:
            headers: List of column header names
            rows: List of rows, each row is a list of values
        """
        self.headers = headers
        self.original_data = rows
        self._populate_table(rows)

    def _populate_table(self, rows: List[List[Any]]):
        """
        Populate table with data.

        Args:
            rows: List of rows to display
        """
        self.table.clear()
        self.table.setRowCount(len(rows))
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)

        for row_idx, row_data in enumerate(rows):
            for col_idx, cell_value in enumerate(row_data):
                item = QTableWidgetItem(str(cell_value))
                # Enable sorting by setting data with proper type
                if isinstance(cell_value, (int, float)):
                    item.setData(Qt.ItemDataRole.UserRole, cell_value)
                self.table.setItem(row_idx, col_idx, item)

        self.update_row_count(len(rows))

    def apply_filter(self):
        """Filter rows based on search text."""
        search_text = self.search_input.text().lower()

        if not search_text:
            self._populate_table(self.original_data)
            return

        filtered_rows = [
            row
            for row in self.original_data
            if any(search_text in str(cell).lower() for cell in row)
        ]
        self._populate_table(filtered_rows)

    def clear_filter(self):
        """Clear search filter and show all rows."""
        self.search_input.clear()
        self._populate_table(self.original_data)

    def copy_selected(self):
        """Copy selected rows to clipboard."""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return

        rows_data = []
        for range_ in selected_ranges:
            for row in range(range_.topRow(), range_.bottomRow() + 1):
                row_data = []
                for col in range(range_.leftColumn(), range_.rightColumn() + 1):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                rows_data.append("\t".join(row_data))

        clipboard_text = "\n".join(rows_data)
        QApplication.clipboard().setText(clipboard_text)

    def export_to_csv(self):
        """Export table data to CSV file."""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", "", "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(self.headers)
                writer.writerows(self.original_data)
        except Exception as e:
            print(f"Error exporting to CSV: {e}")

    def export_to_excel(self, file_path: str):
        """
        Export table data to Excel file.

        Args:
            file_path: Path to save Excel file

        Raises:
            ImportError: If openpyxl is not installed
            Exception: If export fails
        """
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            # Write headers
            for col_idx, header in enumerate(self.headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data
            for row_idx, row_data in enumerate(self.original_data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            wb.save(file_path)
        except ImportError:
            print("openpyxl is not installed. Install it with: pip install openpyxl")
        except Exception as e:
            print(f"Error exporting to Excel: {e}")

    def show_context_menu(self, position):
        """Show context menu on right-click."""
        menu = QMenu(self)

        copy_action = QAction("Copy Selected", self)
        copy_action.triggered.connect(self.copy_selected)
        menu.addAction(copy_action)

        export_csv_action = QAction("Export to CSV", self)
        export_csv_action.triggered.connect(self.export_to_csv)
        menu.addAction(export_csv_action)

        export_excel_action = QAction("Export to Excel", self)

        def export_excel_callback():
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export to Excel", "", "Excel Files (*.xlsx);;All Files (*)"
            )
            if file_path:
                self.export_to_excel(file_path)

        export_excel_action.triggered.connect(export_excel_callback)
        menu.addAction(export_excel_action)

        menu.exec(self.table.viewport().mapToGlobal(position))

    def update_row_count(self, count: int):
        """
        Update row count label.

        Args:
            count: Number of rows currently displayed
        """
        self.row_count_label.setText(f"Rows: {count:,}")

    def get_selected_rows(self) -> List[List[Any]]:
        """
        Get selected rows as list of lists.

        Returns:
            List of selected rows
        """
        selected_rows = []
        selected_ranges = self.table.selectedRanges()

        for range_ in selected_ranges:
            for row in range(range_.topRow(), range_.bottomRow() + 1):
                row_data = []
                for col in range(len(self.headers)):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                selected_rows.append(row_data)

        return selected_rows

    def get_all_data(self) -> Dict[str, Any]:
        """
        Get all table data as dictionary.

        Returns:
            Dictionary with 'headers' and 'rows' keys
        """
        return {"headers": self.headers, "rows": self.original_data}
