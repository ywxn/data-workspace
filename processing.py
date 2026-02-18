"""
Data loading and processing utilities.

Handles data import from various sources (database, CSV, Excel)
using SQL-first workflows and minimal in-memory processing.
"""

import csv
import os
import re
import sqlite3
import tempfile
from typing import Any, Dict, List, Tuple, Optional

from connector import DatabaseConnector
from logger import get_logger
from security_validators import validate_sql_security
from constants import (
    DB_MAX_ROWS_IN_MEMORY,
    DB_READ_CHUNK_SIZE,
    SQL_LARGE_TYPES,
    SAMPLE_ROWS_DEFAULT,
    MERGE_MAX_ESTIMATED_ROWS,
    MERGE_MAX_ROW_MULTIPLIER,
)

logger = get_logger(__name__)


def _quote_identifier(connector: DatabaseConnector, name: str) -> str:
    preparer = connector.engine.dialect.identifier_preparer
    parts = name.split(".")
    return ".".join(preparer.quote(part) for part in parts)


def _quote_sqlite_identifier(name: str) -> str:
    return f"\"{name.replace('"', '""')}\""


def _normalize_columns(raw_columns: List[Any]) -> List[str]:
    columns: List[str] = []
    seen: Dict[str, int] = {}
    for idx, col in enumerate(raw_columns, 1):
        base = str(col).strip() if col not in (None, "") else f"column_{idx}"
        base = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_") or f"column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        name = f"{base}_{count}" if count > 1 else base
        columns.append(name)
    return columns


def _safe_table_name(file_path: str, existing: Optional[set] = None) -> str:
    base = os.path.splitext(os.path.basename(file_path))[0]
    name = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_") or "table"
    if name[0].isdigit():
        name = f"t_{name}"
    existing = existing or set()
    candidate = name
    suffix = 1
    while candidate in existing:
        suffix += 1
        candidate = f"{name}_{suffix}"
    existing.add(candidate)
    return candidate


def _get_columns_with_types(
    connector: DatabaseConnector, table_name: str
) -> List[Dict[str, Any]]:
    from sqlalchemy import inspect

    inspector = inspect(connector.engine)
    return inspector.get_columns(table_name)


def _filter_large_columns(columns: List[Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    safe_cols: List[str] = []
    skipped: List[str] = []

    for col in columns:
        col_name = col.get("name")
        col_type = str(col.get("type", "")).lower()
        if any(t in col_type for t in SQL_LARGE_TYPES):
            skipped.append(col_name)
            continue
        safe_cols.append(col_name)

    return safe_cols, skipped


def _execute_query(
    connector: DatabaseConnector,
    query: str,
    params: Optional[Dict[str, Any]] = None,
    max_rows: Optional[int] = None,
    chunksize: int = DB_READ_CHUNK_SIZE,
) -> Tuple[Dict[str, Any], bool]:
    if not connector.connection:
        raise RuntimeError("No active database connection.")

    is_safe, error_msg = validate_sql_security(query, params)
    if not is_safe:
        raise RuntimeError(error_msg)

    from sqlalchemy import text

    result = connector.connection.execute(text(query), params or {})
    columns = list(result.keys())
    rows: List[Tuple[Any, ...]] = []
    truncated = False

    if max_rows is None:
        rows = result.fetchall()
        return {"columns": columns, "rows": rows}, False

    while True:
        chunk = result.fetchmany(chunksize)
        if not chunk:
            break
        rows.extend(chunk)
        if len(rows) >= max_rows:
            rows = rows[:max_rows]
            truncated = True
            break

    return {"columns": columns, "rows": rows}, truncated


def _rows_to_dicts(columns: List[str], rows: List[Tuple[Any, ...]]) -> List[Dict[str, Any]]:
    return [dict(zip(columns, row)) for row in rows]


def _estimate_table_rows(
    connector: DatabaseConnector, table_name: str
) -> Optional[int]:
    if not connector.connection:
        return None

    quoted_table = _quote_identifier(connector, table_name)
    query = f"SELECT COUNT(1) AS row_count FROM {quoted_table}"
    is_safe, error_msg = validate_sql_security(query)
    if not is_safe:
        logger.warning(error_msg)
        return None

    try:
        from sqlalchemy import text

        result = connector.connection.execute(text(query))
        row = result.fetchone()
        return int(row[0]) if row else None
    except Exception as e:
        logger.warning(f"Row count estimate failed for {table_name}: {str(e)}")
        return None


def _load_table_sample(
    connector: DatabaseConnector, table_name: str, limit: int = SAMPLE_ROWS_DEFAULT
) -> Tuple[List[Dict[str, Any]], List[str]]:
    columns = _get_columns_with_types(connector, table_name)
    safe_cols, skipped_cols = _filter_large_columns(columns)

    if not safe_cols:
        return [], skipped_cols

    quoted_table = _quote_identifier(connector, table_name)
    quoted_cols = ", ".join(
        f"{quoted_table}.{_quote_identifier(connector, col)}" for col in safe_cols
    )
    query = f"SELECT {quoted_cols} FROM {quoted_table} LIMIT {limit}"

    try:
        result, _ = _execute_query(connector, query, max_rows=limit)
        return _rows_to_dicts(result["columns"], result["rows"]), skipped_cols
    except Exception as e:
        logger.warning(f"Sample load failed for {table_name}: {str(e)}")
        return [], skipped_cols


def _collect_table_info(
    connector: DatabaseConnector, table_name: str
) -> Tuple[Dict[str, Any], List[str]]:
    columns = _get_columns_with_types(connector, table_name)
    safe_cols, skipped_cols = _filter_large_columns(columns)

    column_types = {col.get("name"): str(col.get("type", "")) for col in columns}
    row_count = _estimate_table_rows(connector, table_name) or 0
    sample_rows, _ = _load_table_sample(connector, table_name, SAMPLE_ROWS_DEFAULT)

    return (
        {
            "columns": safe_cols,
            "column_types": column_types,
            "row_count": row_count,
            "sample_rows": sample_rows,
        },
        skipped_cols,
    )


def _find_join_condition(
    inspector,
    left_table: str,
    right_table: str,
) -> Optional[Tuple[List[str], List[str], str, str]]:
    for fk in inspector.get_foreign_keys(left_table):
        if fk.get("referred_table") == right_table:
            return (
                fk.get("constrained_columns", []),
                fk.get("referred_columns", []),
                left_table,
                right_table,
            )

    for fk in inspector.get_foreign_keys(right_table):
        if fk.get("referred_table") == left_table:
            return (
                fk.get("referred_columns", []),
                fk.get("constrained_columns", []),
                left_table,
                right_table,
            )

    return None


def _estimate_join_fanout(
    connector: DatabaseConnector,
    left_table: str,
    right_table: str,
    left_cols: List[str],
    right_cols: List[str],
    sample_size: int = 10000,
) -> Optional[float]:
    """
    Estimate fanout factor of a join using key duplication sampling.
    Returns multiplicative fanout (>=1).
    """
    try:
        left_q = _quote_identifier(connector, left_table)
        right_q = _quote_identifier(connector, right_table)

        left_cols_q = ", ".join(_quote_identifier(connector, c) for c in left_cols)
        right_cols_q = ", ".join(_quote_identifier(connector, c) for c in right_cols)

        left_sql = f"SELECT {left_cols_q} FROM {left_q} LIMIT {sample_size}"
        right_sql = f"SELECT {right_cols_q} FROM {right_q} LIMIT {sample_size}"

        left_result, _ = _execute_query(connector, left_sql, max_rows=sample_size)
        right_result, _ = _execute_query(connector, right_sql, max_rows=sample_size)

        left_rows = left_result["rows"]
        right_rows = right_result["rows"]

        if not left_rows or not right_rows:
            return 1.0

        left_unique = len({tuple(row) for row in left_rows})
        right_unique = len({tuple(row) for row in right_rows})

        left_count = len(left_rows)
        right_count = len(right_rows)

        left_dup_rate = 1 - (left_unique / left_count) if left_count else 0
        right_dup_rate = 1 - (right_unique / right_count) if right_count else 0

        def rate_to_fanout(rate: float) -> float:
            if rate <= 0:
                return 1.0
            return 1.0 / (1.0 - rate)

        fanout = max(rate_to_fanout(left_dup_rate), rate_to_fanout(right_dup_rate))
        return max(1.0, fanout)

    except Exception as e:
        logger.debug(f"Fanout estimate failed {left_table}->{right_table}: {e}")
        return None

def _find_best_join_path(
    connector: DatabaseConnector,
    fk_graph: Dict[str, List[Tuple[str, dict]]],
    start: str,
    target: str,
    max_depth: int = 4,
):
    """
    Enumerate FK paths and choose lowest fan-out path.
    """

    from collections import deque

    best_path = None
    best_score = float("inf")

    queue = deque([(start, [], 1.0)])  # (table, path, cumulative_fanout)

    while queue:
        table, path, fanout = queue.popleft()

        if len(path) > max_depth:
            continue

        if table == target:
            if fanout < best_score:
                best_score = fanout
                best_path = path
            continue

        for neighbor, fk in fk_graph.get(table, []):
            if any(step[1] == neighbor for step in path):
                continue

            left_cols = fk.get("constrained_columns", [])
            right_cols = fk.get("referred_columns", [])

            step_fanout = _estimate_join_fanout(
                connector,
                table,
                neighbor,
                left_cols,
                right_cols,
            ) or 1.0

            queue.append(
                (
                    neighbor,
                    path + [(table, neighbor, fk)],
                    fanout * step_fanout,
                )
            )

    return best_path

def _build_join_query(
    connector: DatabaseConnector, table_names: List[str]
) -> Tuple[Optional[str], Dict[str, List[str]], List[str]]:
    """
    Build a safe SQL join query across multiple tables.

    Supports:
    - multi-hop FK joins
    - arbitrary join order
    - missing FK constraints (name-based inference)
    - row-count guard
    - fan-out explosion detection
    """

    from sqlalchemy import inspect
    from collections import deque

    inspector = inspect(connector.engine)

    # ---------- collect safe columns ----------
    safe_cols_by_table: Dict[str, List[str]] = {}
    skipped_cols: List[str] = []

    for table_name in table_names:
        columns = _get_columns_with_types(connector, table_name)
        safe_cols, skipped = _filter_large_columns(columns)
        safe_cols_by_table[table_name] = safe_cols
        skipped_cols.extend([f"{table_name}.{col}" for col in skipped])

    # ---------- base row count ----------
    row_counts: Dict[str, Optional[int]] = {
        t: _estimate_table_rows(connector, t) for t in table_names
    }
    known_counts = [c for c in row_counts.values() if c is not None]
    base_rows = max(known_counts) if known_counts else None
    cumulative_fanout = 1.0

    # ---------- build FK graph ----------
    fk_graph: Dict[str, List[Tuple[str, dict]]] = {}
    for table in inspector.get_table_names():
        fk_graph.setdefault(table, [])
        for fk in inspector.get_foreign_keys(table):
            ref = fk.get("referred_table")
            if not ref:
                continue
            fk_graph[table].append((ref, fk))
            fk_graph.setdefault(ref, [])

    # ---------- helpers ----------
    def find_path(start: str, target: str):
        queue = deque([(start, [])])
        visited = set()

        while queue:
            table, path = queue.popleft()
            if table == target:
                return path
            if table in visited:
                continue
            visited.add(table)

            for neighbor, fk in fk_graph.get(table, []):
                queue.append((neighbor, path + [(table, neighbor, fk)]))

        return None

    def infer_join(left: str, right: str):
        cols_left = {c["name"] for c in _get_columns_with_types(connector, left)}
        cols_right = {c["name"] for c in _get_columns_with_types(connector, right)}
        common = cols_left & cols_right

        candidates = [
            c for c in common if c.lower().endswith("_id") or c.lower() == "id"
        ]
        if not candidates:
            return None

        col = candidates[0]
        return left, right, [col], [col]

    def check_fanout(left_table, right_table, left_cols, right_cols):
        nonlocal cumulative_fanout

        fanout = _estimate_join_fanout(
            connector,
            left_table,
            right_table,
            left_cols,
            right_cols,
        )

        if fanout:
            cumulative_fanout *= fanout

            if base_rows:
                estimated_rows = base_rows * cumulative_fanout

                if (
                    estimated_rows > MERGE_MAX_ESTIMATED_ROWS
                    and estimated_rows > base_rows * MERGE_MAX_ROW_MULTIPLIER
                ):
                    logger.warning(
                        "Join aborted (fanout explosion): "
                        f"{left_table}->{right_table} "
                        f"fanout≈{fanout:.2f}, "
                        f"est≈{estimated_rows:,.0f} rows"
                    )
                    return False
        return True

    # ---------- build joins ----------
    base_table = table_names[0]
    joined = {base_table}
    join_clauses: List[str] = []

    for target in table_names[1:]:
        if target in joined:
            continue

        path = None

        for candidate in joined:
            p = _find_best_join_path(connector, fk_graph, candidate, target)
            if p:
                path = p
                break

        if not path:
            inferred = infer_join(base_table, target)
            if not inferred:
                logger.warning(f"No join path found for {base_table}->{target}")
                return None, safe_cols_by_table, skipped_cols

            left_table, right_table, left_cols, right_cols = inferred
            if not check_fanout(left_table, right_table, left_cols, right_cols):
                return None, safe_cols_by_table, skipped_cols

            left_q = _quote_identifier(connector, left_table)
            right_q = _quote_identifier(connector, right_table)
            conditions = " AND ".join(
                f"{left_q}.{_quote_identifier(connector, l)} = {right_q}.{_quote_identifier(connector, r)}"
                for l, r in zip(left_cols, right_cols)
            )
            join_clauses.append(f"JOIN {right_q} ON {conditions}")
            joined.add(right_table)
            continue

        for left_table, right_table, fk in path:
            if right_table in joined:
                continue

            left_cols = fk.get("constrained_columns", [])
            right_cols = fk.get("referred_columns", [])

            if not left_cols or not right_cols:
                inferred = infer_join(left_table, right_table)
                if not inferred:
                    logger.warning(f"No join columns found for {left_table}->{right_table}")
                    return None, safe_cols_by_table, skipped_cols
                left_table, right_table, left_cols, right_cols = inferred

            if not check_fanout(left_table, right_table, left_cols, right_cols):
                return None, safe_cols_by_table, skipped_cols

            left_q = _quote_identifier(connector, left_table)
            right_q = _quote_identifier(connector, right_table)
            conditions = " AND ".join(
                f"{left_q}.{_quote_identifier(connector, l)} = {right_q}.{_quote_identifier(connector, r)}"
                for l, r in zip(left_cols, right_cols)
            )
            join_clauses.append(f"JOIN {right_q} ON {conditions}")
            joined.add(right_table)

    select_cols: List[str] = []
    for table_name in table_names:
        table_q = _quote_identifier(connector, table_name)
        for col in safe_cols_by_table.get(table_name, []):
            col_q = _quote_identifier(connector, col)
            alias = f"{table_name}__{col}"
            alias_q = _quote_identifier(connector, alias)
            select_cols.append(f"{table_q}.{col_q} AS {alias_q}")

    if not select_cols:
        return None, safe_cols_by_table, skipped_cols

    base_q = _quote_identifier(connector, base_table)
    query = f"SELECT {', '.join(select_cols)} FROM {base_q}"
    if join_clauses:
        query = f"{query} {' '.join(join_clauses)}"

    return query, safe_cols_by_table, skipped_cols


def _read_csv_rows(file_path: str) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            return [], []
        columns = _normalize_columns(header)
        rows: List[Tuple[Any, ...]] = []
        for row in reader:
            values = list(row)
            if len(values) < len(columns):
                values.extend([""] * (len(columns) - len(values)))
            rows.append(tuple(values[: len(columns)]))
        return columns, rows


def _read_excel_rows(
    file_path: str, sheet_name: Optional[str] = None
) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return [], []

    columns = _normalize_columns(list(header))
    rows: List[Tuple[Any, ...]] = []
    for row in rows_iter:
        values = ["" if value is None else str(value) for value in row]
        if len(values) < len(columns):
            values.extend([""] * (len(columns) - len(values)))
        rows.append(tuple(values[: len(columns)]))
    return columns, rows


def _load_files_into_sqlite(
    file_paths: List[str], sheet_name: Optional[str] = None
) -> Tuple[str, List[str], str]:
    if not file_paths:
        raise ValueError("No file paths provided.")

    temp_db = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    temp_db.close()

    conn = sqlite3.connect(temp_db.name)
    created_tables: List[str] = []
    existing = set()

    try:
        for file_path in file_paths:
            table_name = _safe_table_name(file_path, existing)
            if file_path.lower().endswith(".csv"):
                columns, rows = _read_csv_rows(file_path)
            elif file_path.lower().endswith((".xlsx", ".xls")):
                columns, rows = _read_excel_rows(file_path, sheet_name)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            if not columns:
                raise ValueError(f"No columns detected in file: {file_path}")

            col_defs = ", ".join(
                f"{_quote_sqlite_identifier(col)} TEXT" for col in columns
            )
            conn.execute(
                f"CREATE TABLE {_quote_sqlite_identifier(table_name)} ({col_defs})"
            )

            placeholders = ", ".join(["?"] * len(columns))
            col_list = ", ".join(_quote_sqlite_identifier(col) for col in columns)
            insert_sql = (
                f"INSERT INTO {_quote_sqlite_identifier(table_name)} "
                f"({col_list}) VALUES ({placeholders})"
            )
            if rows:
                conn.executemany(insert_sql, rows)

            created_tables.append(table_name)

        conn.commit()
    finally:
        conn.close()

    return temp_db.name, created_tables, "Files loaded into SQLite workspace"


def load_data(
    source_type: str, source_config: Dict[str, Any]
) -> Tuple[Optional[Dict[str, Any]], str]:
    """
    Load data from a database, CSV, or Excel file into a SQL-native context.

    Args:
        source_type: Type of source ('database', 'csv', 'excel', or 'file')
        source_config: Configuration dictionary containing:
            - For database: {
                'db_type': str,
                'credentials': Dict,
                'table': str | List[str],
                ...
              }
            - For CSV: {'file_path': str}
            - For Excel: {'file_path': str, 'sheet_name': str (optional)}
            - For file: {'file_paths': List[str]}

    Returns:
        Tuple of (context or None, status_message: str)
    """
    try:
        source_type = source_type.lower()
        logger.info(f"Loading data from {source_type} source")

        if source_type == "database":
            return _load_from_database(source_config)
        if source_type == "file":
            return _load_from_files(source_config)
        if source_type == "csv":
            return _load_from_files({"file_paths": [source_config.get("file_path", "")]})
        if source_type == "excel":
            config = {
                "file_paths": [source_config.get("file_path", "")],
                "sheet_name": source_config.get("sheet_name"),
            }
            return _load_from_files(config)

        msg = f"Unsupported source type: {source_type}"
        logger.error(msg)
        return None, msg

    except Exception as e:
        msg = f"Error loading data: {str(e)}"
        logger.error(msg, exc_info=True)
        return None, msg


def _load_from_database(config: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], str]:
    table = config.get("table")
    if not table:
        return None, "Error: 'table' field is required for database source"

    db_type = config.get("db_type")
    credentials = config.get("credentials", {})

    logger.info(f"Connecting to {db_type} database...")
    connector = DatabaseConnector()
    success, message = connector.connect(db_type, credentials)

    if not success:
        return None, message

    try:
        tables = table if isinstance(table, list) else [table]
        table_info: Dict[str, Any] = {}
        skipped_cols_by_table: Dict[str, List[str]] = {}

        for table_name in tables:
            info, skipped_cols = _collect_table_info(connector, table_name)
            table_info[table_name] = info
            if skipped_cols:
                skipped_cols_by_table[table_name] = skipped_cols

        context = {
            "source_type": "database",
            "db_type": db_type,
            "credentials": credentials,
            "tables": tables,
            "table_info": table_info,
            "skipped_columns": skipped_cols_by_table,
        }

        return context, "Schema loaded successfully"
    finally:
        connector.close()


def _load_from_files(config: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], str]:
    file_paths = config.get("file_paths") or []
    if not file_paths:
        return None, "Error: 'file_paths' is required for file source"

    sheet_name = config.get("sheet_name")

    try:
        db_path, tables, message = _load_files_into_sqlite(file_paths, sheet_name)
        connector = DatabaseConnector()
        success, conn_msg = connector.connect("sqlite", {"database": db_path})
        if not success:
            return None, conn_msg

        try:
            table_info: Dict[str, Any] = {}
            skipped_cols_by_table: Dict[str, List[str]] = {}
            for table_name in tables:
                info, skipped_cols = _collect_table_info(connector, table_name)
                table_info[table_name] = info
                if skipped_cols:
                    skipped_cols_by_table[table_name] = skipped_cols

            context = {
                "source_type": "file",
                "db_type": "sqlite",
                "credentials": {"database": db_path},
                "tables": tables,
                "table_info": table_info,
                "file_paths": file_paths,
                "skipped_columns": skipped_cols_by_table,
            }

            return context, message
        finally:
            connector.close()
    except Exception as e:
        logger.error(f"Error loading files: {str(e)}", exc_info=True)
        return None, f"Error loading files: {str(e)}"


def add_files_to_sqlite(
    context: Dict[str, Any], file_paths: List[str]
) -> Tuple[Optional[Dict[str, Any]], str]:
    if not context or context.get("db_type") != "sqlite":
        return None, "Current context is not a SQLite workspace."

    db_path = context.get("credentials", {}).get("database")
    if not db_path:
        return None, "Missing SQLite database path in context."

    existing_list = context.get("tables", [])
    existing_tables = set(existing_list)

    conn = sqlite3.connect(db_path)
    created_tables: List[str] = []
    try:
        for file_path in file_paths:
            table_name = _safe_table_name(file_path, existing_tables)
            if file_path.lower().endswith(".csv"):
                columns, rows = _read_csv_rows(file_path)
            elif file_path.lower().endswith((".xlsx", ".xls")):
                columns, rows = _read_excel_rows(file_path, None)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            if not columns:
                raise ValueError(f"No columns detected in file: {file_path}")

            col_defs = ", ".join(
                f"{_quote_sqlite_identifier(col)} TEXT" for col in columns
            )
            conn.execute(
                f"CREATE TABLE {_quote_sqlite_identifier(table_name)} ({col_defs})"
            )

            placeholders = ", ".join(["?"] * len(columns))
            col_list = ", ".join(_quote_sqlite_identifier(col) for col in columns)
            insert_sql = (
                f"INSERT INTO {_quote_sqlite_identifier(table_name)} "
                f"({col_list}) VALUES ({placeholders})"
            )
            if rows:
                conn.executemany(insert_sql, rows)

            created_tables.append(table_name)

        conn.commit()
    finally:
        conn.close()

    if not created_tables:
        return None, "No new tables were added."

    connector = DatabaseConnector()
    success, message = connector.connect("sqlite", {"database": db_path})
    if not success:
        return None, message

    try:
        table_info = context.get("table_info", {})
        skipped_cols_by_table = context.get("skipped_columns", {})

        for table_name in created_tables:
            info, skipped_cols = _collect_table_info(connector, table_name)
            table_info[table_name] = info
            if skipped_cols:
                skipped_cols_by_table[table_name] = skipped_cols

        updated = dict(context)
        updated["tables"] = existing_list + created_tables
        updated["table_info"] = table_info
        updated["skipped_columns"] = skipped_cols_by_table
        updated["file_paths"] = list(set(context.get("file_paths", [])) | set(file_paths))

        return updated, f"Added {len(created_tables)} table(s) to SQLite workspace"
    finally:
        connector.close()
